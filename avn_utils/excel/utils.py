import openpyxl
from openpyxl import load_workbook
import pandas as pd

def get_df_by_range(excel_file: str
        , excel_sheet: str, excel_range: str
        , excel_header: str) -> pd.DataFrame:
    workbook = load_workbook(excel_file)
    sheet = workbook[excel_sheet]
    #workbook = load_workbook('db\\jee-mains.xlsx')
    #sheet = workbook['jee-mains']

    # Specify range (e.g., A1:C10)
    data = sheet[excel_range] #Range by Single Row
    #data = sheet['A1:C10'] #Range by Row, Column
    #data = sheet['A1:C1'] #Range by Single Row
    #data = sheet['A1:A10'] #Range by Single Column
    rows = [[cell.value for cell in row] for row in data]
    columns = [cell.value for cell in sheet[excel_header][0]]
    #columns = [cell.value for cell in sheet['A1:C1'][0]] #Header by multi column
    #columns = [cell.value for cell in sheet['A1:A1'][0]] #Header by single column
    df = pd.DataFrame(rows, columns=columns)
    return df
'''Example Usage
excel_file = 'db\\jee-mains.xlsx'
excel_sheet = 'jee-mains'
excel_range = 'A2:H10'  # Example range
excel_header = 'A1:H1'  # Example header range
print(get_df_by_range(excel_file, excel_sheet
        , excel_range, excel_header))
'''

def get_df_by_cell_ids(excel_file, sheet_name, cell_ids, header_cell_ids):
    # Validate that the number of cell_ids matches the number of header_cell_ids
    if len(cell_ids) != len(header_cell_ids):
        raise ValueError("The number of cell IDs must match the number of header cell IDs")
    
    # Load the Excel workbook and select the specified sheet
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheet_name]
    
    # Create a list to store values corresponding to provided cell_ids
    values = []
    # Create a list to store header values from header_cell_ids
    headers = []
    
    # Iterate through header_cell_ids to fetch header values
    for header_id in header_cell_ids:
        # Get value from the header cell, or use the cell ID if empty/invalid
        header_value = sheet[header_id].value if sheet[header_id].value is not None else header_id
        headers.append(header_value)
    
    # Iterate through cell_ids to fetch data values
    for cell_id in cell_ids:
        # Get value from the cell, or None if cell is empty/invalid
        value = sheet[cell_id].value if sheet[cell_id].value is not None else None
        values.append(value)
    
    # Create DataFrame with header values as columns and a single row of values
    df = pd.DataFrame([values], columns=headers)
    
    # Close the workbook
    workbook.close()
    
    return df

'''# Example usage
excel_file = 'db\\jee-mains.xlsx'  # Replace with your Excel file path
sheet_name = 'jee-mains'        # Replace with your sheet name
cell_ids = ['A2', 'B2', 'D4', 'F6']
header_cell_ids = ['A1', 'B1', 'D1', 'F1']
result_df = get_df_by_cell_ids(excel_file, sheet_name, cell_ids, header_cell_ids)
print(result_df)
'''